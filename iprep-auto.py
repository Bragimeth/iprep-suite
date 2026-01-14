import requests
import concurrent.futures
import time
import os
import urllib3
from rich.progress import Progress
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.align import Align
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pyfiglet import Figlet

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ================== API KEYS ==================
ABUSEIPDB_KEY = "ABUSEIPDB_KEY"
IPQS_KEY = "IPQS_KEY"
IPINFO_KEY = "IPINFO_KEY"

# ================== TARGET (SIEM_NAME) ==================
TARGET_URL = "SIEM_API_ENDPOINT"
WATCHLIST_ID = "SIEM_WATCHLIST_ID"
CUST_ID = "SIEM_CUSTOMER_ID"
TARGET_AUTH = "SIEM_AUTH_TOKEN"
TARGET_COOKIE = "SIEM_SESSION"

# ================== AYARLAR ==================
IP_FILE = "ip.txt"
ABUSE_THRESHOLD = 20
TIMEOUT = 5
MAX_WORKERS = 10

# ================== URLS ==================
ABUSE_URL = "https://api.abuseipdb.com/api/v2/check"
IPQS_URL = "https://ipqualityscore.com/api/json/ip/{api_key}/{ip}"
IPINFO_URL = "https://ipinfo.io/{ip}/json?token={api_key}"

HEADERS_ABUSE = {"Key": ABUSEIPDB_KEY, "Accept": "application/json"}

# ================== EXCEL RENKLER ==================
RED = PatternFill(start_color="FF4C4C", end_color="FF4C4C", fill_type="solid")
YELLOW = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
GREEN = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
BLUE = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")

# ================== FUNCTIONS ==================
def check_abuseipdb(ip):
    try:
        r = requests.get(
            ABUSE_URL,
            headers=HEADERS_ABUSE,
            params={"ipAddress": ip, "maxAgeInDays": 90},
            timeout=TIMEOUT
        )
        return r.json()["data"]["abuseConfidenceScore"]
    except:
        return None


def check_ipqs(ip):
    try:
        r = requests.get(IPQS_URL.format(api_key=IPQS_KEY, ip=ip), timeout=TIMEOUT)
        d = r.json()
        return d.get("fraud_score"), d.get("proxy"), d.get("tor"), d.get("bot")
    except:
        return None, None, None, None


def get_owner(ip):
    try:
        r = requests.get(IPINFO_URL.format(ip=ip, api_key=IPINFO_KEY), timeout=TIMEOUT)
        return r.json().get("org", "Unknown")
    except:
        return "Unknown"


def send_to_target_whitelist(ip):
    payload = [{
        "entryValue": ip,
        "state": "Enabled",
        "description": "Auto Whitelist"
    }]

    headers = {
        "Content-Type": "application/json",
        "Authorization": TARGET_AUTH,
        "Cookie": TARGET_COOKIE
    }

    try:
        r = requests.post(
            TARGET_URL,
            params={"watchlistId": WATCHLIST_ID, "custId": CUST_ID},
            json=payload,
            headers=headers,
            timeout=TIMEOUT,
            verify=False
        )

        resp = r.json() if r.content else {}

        if "Duplicate entry value" in str(resp):
            return "Zaten Var"
        if r.status_code in (200, 201) and resp.get("status") != "Failed":
            return "Evet"

        return "Hata"
    except:
        return "Hata"


def analyze_ip(ip):
    abuse = check_abuseipdb(ip)
    fraud, proxy, tor, bot = check_ipqs(ip)
    owner = get_owner(ip)

    if abuse is not None and abuse >= ABUSE_THRESHOLD:
        status = "BLACKLIST"
        sent = "Hayır"
    else:
        status = "WHITELIST"
        sent = send_to_target_whitelist(ip)

    return [
        ip,
        abuse,
        fraud,
        "Evet" if proxy else "Hayır",
        "Evet" if tor else "Hayır",
        "Evet" if bot else "Hayır",
        status,
        sent,
        owner
    ]


def apply_color(cell, score):
    if score is None:
        return
    if score >= 20:
        cell.fill = RED
    elif score >= 5:
        cell.fill = YELLOW
    else:
        cell.fill = GREEN


def apply_sent_color(cell, sent):
    if sent == "Evet":
        cell.fill = GREEN
    elif sent == "Zaten Var":
        cell.fill = BLUE
    elif sent == "Hata":
        cell.fill = RED


# ================== MAIN ==================
def main():
    console = Console()

    banner = Figlet(font="big").renderText("0xLdap")
    console.print(
        Panel(
            Align.center(banner),
            title="IP Reputation Scanner",
            subtitle=datetime.now().strftime("%Y-%m-%d %H:%M")
        )
    )

    console.print(
        Align.center("[dim]Author: Anıl Doğan | Cyber Security Specialist[/dim]\n")
    )

    with open(IP_FILE) as f:
        ips = [i.strip() for i in f if i.strip()]

    wb = Workbook()
    ws_black = wb.active
    ws_black.title = "Blacklist"
    ws_white = wb.create_sheet("Whitelist")

    headers = ["IP", "AbuseIPDB", "IPQS Fraud", "Proxy", "Tor", "Bot",
               "Result", "SentToTarget", "Owner"]
    ws_black.append(headers)
    ws_white.append(headers)

    results = []
    sent_ok = sent_dup = sent_err = 0
    start = time.time()

    with Progress() as progress:
        task = progress.add_task("[cyan]Checking IPs...", total=len(ips))
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for r in executor.map(analyze_ip, ips):
                results.append(r)
                progress.update(task, advance=1)

    table = Table(title="IP Reputation Results")
    for h in headers:
        table.add_column(h)

    for r in results:
        ws = ws_black if r[6] == "BLACKLIST" else ws_white
        ws.append(r)

        apply_color(ws.cell(row=ws.max_row, column=2), r[1])
        apply_sent_color(ws.cell(row=ws.max_row, column=8), r[7])

        if r[7] == "Evet":
            sent_ok += 1
        elif r[7] == "Zaten Var":
            sent_dup += 1
        elif r[7] == "Hata":
            sent_err += 1

        color = "red" if r[6] == "BLACKLIST" else "green"
        table.add_row(
            r[0], str(r[1]), str(r[2]), r[3], r[4], r[5],
            f"[{color}]{r[6]}[/{color}]", r[7], r[8]
        )

    console.print(table)

    filename = f"ip_reputation_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)

    console.print(f"\n✅ Rapor: [bold yellow]{filename}[/bold yellow]")
    console.print(f"🔢 IP sayısı: {len(ips)}")
    console.print(f"⏱ Süre: {round(time.time() - start, 2)} sn")

    console.print("\n📤 SIEM / SOAR ENTEGRASYONU")
    console.print(f"✅ Eklenen IP: [bold green]{sent_ok}[/bold green]")
    console.print(f"⚠ Zaten Var: [bold blue]{sent_dup}[/bold blue]")
    console.print(f"❌ Hata alan IP: [bold red]{sent_err}[/bold red]")


if __name__ == "__main__":
    main()
